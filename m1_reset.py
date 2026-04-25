import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

FILE_PATH = "data.xlsx"

COLUMNS = [
    "企業名",
    "再評価フラグ",
    "業種_AI",
    "業種_補正",
    "安定性_AI",
    "安定性_補正",
    "年収_AI",
    "年収_補正",
    "成長性_AI",
    "成長性_補正",
    "WLB_AI",
    "WLB_補正",
    "総合点",
    "判定",
    "企業評価コメント",
    "システム実行コメント",
    "更新日時"
]

def create_template_excel(file_path: str) -> None:
    # ★ 既存でも問答無用で作り直す
    df = pd.DataFrame(columns=COLUMNS)
    df.to_excel(file_path, index=False)

    wb = load_workbook(file_path)
    ws = wb.active
    ws.title = "企業評価"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(file_path)
    print(f"台帳を作り直しました: {file_path}")

if __name__ == "__main__":
    create_template_excel(FILE_PATH)